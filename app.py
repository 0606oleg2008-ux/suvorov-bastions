from flask import Flask, request, jsonify, render_template, send_file
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
DATA_FILE = 'data.json'
EXCEL_FILE = 'data.xlsx'

if not os.path.exists(DATA_FILE):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump({"frozen": False, "mapBackground": "", "housesPositions": [], "roadPoints": []}, f)

def load_data():
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {"frozen": False, "mapBackground": "", "housesPositions": [], "roadPoints": []}

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_status_text(status):
    map_status = {
        'not_started': 'Не начат',
        'in_progress': 'В процессе',
        'completed': 'Завершён'
    }
    return map_status.get(status, 'Неизвестно')

def get_status_color(status):
    map_color = {
        'not_started': 'FF6B6B',
        'in_progress': 'FFD93D',
        'completed': '6BCB77'
    }
    return map_color.get(status, 'FFFFFF')

def get_execution_status(floor):
    plan_start = floor.get('plan_start', '')
    plan_end = floor.get('plan_end', '')
    fact_start = floor.get('fact_start', '')
    fact_end = floor.get('fact_end', '')
    today = datetime.now().date()
    
    if not plan_start and not plan_end:
        return 'Неизвестно'
    
    try:
        plan_start_date = datetime.strptime(plan_start, '%Y-%m-%d').date() if plan_start else None
        plan_end_date = datetime.strptime(plan_end, '%Y-%m-%d').date() if plan_end else None
        fact_start_date = datetime.strptime(fact_start, '%Y-%m-%d').date() if fact_start else None
        fact_end_date = datetime.strptime(fact_end, '%Y-%m-%d').date() if fact_end else None
    except:
        return 'Ошибка дат'
    
    if fact_end_date:
        if plan_end_date:
            if fact_end_date <= plan_end_date:
                return '✅ Успеваем'
            else:
                return '⚠️ Отстаём'
        else:
            return '✅ Завершён (без плана)'
    
    if fact_start_date:
        if plan_end_date:
            if fact_start_date > plan_end_date:
                return '🔴 Просрочен'
            elif today > plan_end_date:
                return '🔴 Просрочен'
            else:
                return '⏳ В процессе'
        else:
            return '⏳ В процессе'
    
    if plan_start_date:
        if today > plan_start_date:
            return '🔴 Просрочен'
        else:
            return '⏳ Ожидание'
    
    return 'Неизвестно'

def generate_excel(data):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Строительство'

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border_style = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        number_alignment = Alignment(horizontal="right", vertical="center")
        text_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        even_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        headers = ['Дом', 'Название дома', 'План начала дома', 'План окончания дома',
                   'Факт начала дома', 'Факт окончания дома', 'Всего смен', 'Всего человек',
                   'Этап', 'Статус этапа', 'План начала этапа', 'План окончания этапа',
                   'Факт начала этапа', 'Факт окончания этапа', 'Выполнение',
                   'Смена', 'Дата смены', 'Человек в смене']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        row_num = 2
        house_ids = [k for k in sorted(data.keys(), key=lambda x: int(x) if x.isdigit() else 0) 
                    if k not in ['frozen', 'mapBackground', 'housesPositions', 'roadPoints']]

        for house_id in house_ids:
            house = data[house_id]
            floors = house.get('floors', [])

            total_shifts = 0
            total_workers = 0
            for f in floors:
                shifts = f.get('shifts', [])
                total_shifts += len(shifts)
                for s in shifts:
                    total_workers += s.get('workers', 0)

            fill = even_row_fill if row_num % 2 == 0 else odd_row_fill
            ws.cell(row=row_num, column=1, value=f"#{house_id}").font = Font(bold=True, size=11)
            ws.cell(row=row_num, column=2, value=house.get('name', '')).font = Font(bold=True)
            ws.cell(row=row_num, column=3, value=house.get('plan_start', ''))
            ws.cell(row=row_num, column=4, value=house.get('plan_end', ''))
            ws.cell(row=row_num, column=5, value=house.get('fact_start', ''))
            ws.cell(row=row_num, column=6, value=house.get('fact_end', ''))
            ws.cell(row=row_num, column=7, value=total_shifts).alignment = number_alignment
            ws.cell(row=row_num, column=8, value=total_workers).alignment = number_alignment
            for col in [9,10,11,12,13,14,15,16,17,18]:
                ws.cell(row=row_num, column=col).value = ''
            for col in range(1, len(headers)+1):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).border = border_style
                ws.cell(row=row_num, column=col).alignment = text_alignment
                if col in [3,4,5,6,7,8]:
                    ws.cell(row=row_num, column=col).alignment = number_alignment
            ws.row_dimensions[row_num].outline_level = 0
            row_num += 1

            for floor in floors:
                shifts = floor.get('shifts', [])
                total_shifts_floor = len(shifts)
                total_workers_floor = sum(s.get('workers', 0) for s in shifts)

                fill = even_row_fill if row_num % 2 == 0 else odd_row_fill
                status_val = floor.get('status', '')
                status_text = get_status_text(status_val)
                status_color = get_status_color(status_val)
                status_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

                ws.cell(row=row_num, column=1, value=f"#{house_id}").font = Font(bold=False, size=10)
                ws.cell(row=row_num, column=2, value=house.get('name', '')).font = Font(bold=False)
                ws.cell(row=row_num, column=7, value=total_shifts_floor).alignment = number_alignment
                ws.cell(row=row_num, column=8, value=total_workers_floor).alignment = number_alignment
                ws.cell(row=row_num, column=9, value=floor.get('name', '')).font = Font(bold=True)
                status_cell = ws.cell(row=row_num, column=10, value=status_text)
                status_cell.fill = status_fill
                status_cell.font = Font(bold=True, color="000000")
                ws.cell(row=row_num, column=11, value=floor.get('plan_start', ''))
                ws.cell(row=row_num, column=12, value=floor.get('plan_end', ''))
                ws.cell(row=row_num, column=13, value=floor.get('fact_start', ''))
                ws.cell(row=row_num, column=14, value=floor.get('fact_end', ''))
                exec_status = get_execution_status(floor)
                ws.cell(row=row_num, column=15, value=exec_status)
                for col in [16,17,18]:
                    ws.cell(row=row_num, column=col).value = ''

                for col in range(1, len(headers)+1):
                    if col != 10:
                        ws.cell(row=row_num, column=col).fill = fill
                    ws.cell(row=row_num, column=col).border = border_style
                    ws.cell(row=row_num, column=col).alignment = text_alignment
                    if col in [3,4,5,6,7,8,11,12,13,14,18]:
                        ws.cell(row=row_num, column=col).alignment = number_alignment
                ws.row_dimensions[row_num].outline_level = 1
                row_num += 1

                for shift in shifts:
                    fill = even_row_fill if row_num % 2 == 0 else odd_row_fill
                    ws.cell(row=row_num, column=1, value=f"#{house_id}").font = Font(bold=False, size=9)
                    ws.cell(row=row_num, column=2, value=house.get('name', '')).font = Font(bold=False)
                    ws.cell(row=row_num, column=16, value=shift.get('name', ''))
                    ws.cell(row=row_num, column=17, value=shift.get('date', ''))
                    ws.cell(row=row_num, column=18, value=shift.get('workers', 0)).alignment = number_alignment

                    for col in range(1, len(headers)+1):
                        ws.cell(row=row_num, column=col).fill = fill
                        ws.cell(row=row_num, column=col).border = border_style
                        ws.cell(row=row_num, column=col).alignment = text_alignment
                        if col in [7,8,18]:
                            ws.cell(row=row_num, column=col).alignment = number_alignment
                    ws.row_dimensions[row_num].outline_level = 2
                    row_num += 1

        col_widths = {
            1: 10, 2: 25, 3: 16, 4: 16, 5: 16, 6: 16,
            7: 12, 8: 14, 9: 25, 10: 18, 11: 16, 12: 16,
            13: 16, 14: 16, 15: 20, 16: 20, 17: 16, 18: 16
        }
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.sheet_view.show_outline_symbols = True

        if row_num > 1:
            ws.auto_filter.ref = ws.dimensions

        wb.save(EXCEL_FILE)
        print(f"✅ Excel обновлён: {EXCEL_FILE}")
        return True
    except Exception as e:
        print(f"❌ Ошибка генерации Excel: {e}")
        return False

# ------------------ Маршруты ------------------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/data', methods=['GET'])
def get_data():
    return jsonify(load_data())

@app.route('/api/data', methods=['POST'])
def post_data():
    try:
        new_data = request.json
        save_data(new_data)
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/freeze', methods=['POST'])
def set_freeze():
    try:
        data = load_data()
        data['frozen'] = True
        save_data(data)
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/houses', methods=['GET', 'POST'])
def api_houses():
    if request.method == 'GET':
        data = load_data()
        return jsonify(data.get('housesPositions', []))
    else:
        data = load_data()
        data['housesPositions'] = request.json
        save_data(data)
        return jsonify({'status': 'ok'})

@app.route('/api/road', methods=['GET', 'POST'])
def api_road():
    if request.method == 'GET':
        data = load_data()
        return jsonify(data.get('roadPoints', []))
    else:
        data = load_data()
        data['roadPoints'] = request.json
        save_data(data)
        return jsonify({'status': 'ok'})

@app.route('/api/background', methods=['GET', 'POST'])
def api_background():
    if request.method == 'GET':
        data = load_data()
        return jsonify({'background': data.get('mapBackground', '')})
    else:
        data = load_data()
        data['mapBackground'] = request.json.get('background', '')
        save_data(data)
        return jsonify({'status': 'ok'})

@app.route('/api/excel')
def download_excel():
    data = load_data()
    generate_excel(data)
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True, download_name='Строительство_Суворовские_бастионы.xlsx')
    else:
        return "Файл ещё не создан", 404

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
